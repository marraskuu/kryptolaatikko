from datetime import datetime, timezone
from io import BytesIO

from openpyxl import Workbook
from openpyxl.utils import get_column_letter

from .bitfinex import get_crypto_label
from .portfolio import Portfolio


def _fmt_date(iso: str) -> str:
    dt = datetime.fromisoformat(iso.replace("Z", "+00:00"))
    if dt.tzinfo is None:
        dt = dt.replace(tzinfo=timezone.utc)
    return dt.astimezone().strftime("%d.%m.%Y")


def _fmt_pnl(value: float | None) -> str:
    if value is None:
        return ""
    rounded = round(value, 2)
    text = f"{abs(rounded):.2f}".replace(".", ",")
    if rounded > 0:
        return f"+{text}"
    if rounded < 0:
        return f"-{text}"
    return "0,00"


def _unit_price(trade: dict) -> float:
    amount = trade.get("amount") or 0
    if amount > 0 and trade.get("eurTotal") is not None:
        return trade["eurTotal"] / amount
    return float(trade.get("price") or 0)


def _fmt_unit_price(price: float) -> float:
    if price >= 1000:
        return round(price, 2)
    if price >= 1:
        return round(price, 4)
    if price >= 0.01:
        return round(price, 6)
    return round(price, 8)


def _fmt_quantity(amount: float) -> float:
    if amount >= 1:
        return round(amount, 6)
    if amount >= 0.0001:
        return round(amount, 8)
    return round(amount, 12)


def _sell_profit_loss(trade: dict) -> float:
    profit_loss = trade.get("profitLoss")
    if profit_loss is not None:
        return float(profit_loss)
    if trade.get("profit") is not None:
        return float(trade["profit"])
    cost_basis = trade.get("costBasis", 0)
    return float(trade["eurTotal"]) - float(cost_basis)


def _autosize_columns(ws, widths: list[int]) -> None:
    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width


def build_tax_excel(portfolio_data: dict) -> tuple[BytesIO, str]:
    portfolio = Portfolio(portfolio_data)
    trades = sorted(
        [t for t in portfolio.trades if t["type"] in ("buy", "sell")],
        key=lambda t: t["timestamp"],
    )
    if not trades:
        raise ValueError("Ei kauppoja vientiin.")

    buys = [t for t in trades if t["type"] == "buy"]
    sells = [t for t in trades if t["type"] == "sell"]

    wb = Workbook()
    ws = wb.active
    ws.title = "Kaupat"
    ws.append(
        [
            "Päivämäärä",
            "Krypto",
            "Kappalemäärä",
            "Ostohinta (EUR/kpl)",
            "Myyntihinta (EUR/kpl)",
            "Voitto (+) / Tappio (-) (EUR)",
        ]
    )

    for t in trades:
        label = get_crypto_label(t["symbol"])
        date = _fmt_date(t["timestamp"])
        qty = _fmt_quantity(float(t.get("amount") or 0))
        unit = _fmt_unit_price(_unit_price(t))

        if t["type"] == "buy":
            ws.append([date, label, qty, unit, "", ""])
        else:
            ws.append(
                [
                    date,
                    label,
                    qty,
                    "",
                    unit,
                    _fmt_pnl(_sell_profit_loss(t)),
                ]
            )

    ws_summary = wb.create_sheet("Yhteenveto")
    total_profit = sum(_sell_profit_loss(t) for t in sells)
    realized_by_year = portfolio.realized_profit_by_year()
    tax_by_year = portfolio.tax_owed_by_year()

    ws_summary.append(["Raportin luontipäivä", _fmt_date(datetime.now(timezone.utc).isoformat())])
    ws_summary.append(["Ostoja (kpl)", len(buys)])
    ws_summary.append(["Myyntejä (kpl)", len(sells)])
    ws_summary.append(["Voitot ja tappiot yhteensä, kaikki vuodet (EUR)", _fmt_pnl(total_profit) or "0,00"])
    ws_summary.append(["", ""])
    ws_summary.append(["Verovuosikohtainen erittely (pääomatulovero 30 %)", ""])
    for year in sorted(realized_by_year):
        net = realized_by_year[year]
        ws_summary.append([f"{year} — nettoluovutusvoitto/-tappio (EUR)", _fmt_pnl(net) or "0,00"])
        ws_summary.append(
            [f"{year} — vero 30 % nettovoitosta (EUR, maksat itse)", round(tax_by_year.get(year, 0.0), 2)]
        )
    current_year = datetime.now(timezone.utc).year
    carry = portfolio.loss_carryforward(as_of_year=current_year)
    if carry > 0.01:
        ws_summary.append(["", ""])
        ws_summary.append(
            ["Käyttämätön luovutustappio, siirtyy vähennettäväksi seuraavien 5 v voitoista (EUR)", round(carry, 2)]
        )
    ws_summary.append(["", ""])
    ws_summary.append(
        [
            "Huomautus",
            "Simulaatio Bitfinex-kursseilla. Hinta = EUR/kpl (eurTotal ÷ kappalemäärä). "
            "Vero lasketaan verovuoden NETTOluovutusvoitosta (voitot miinus tappiot) — "
            "Suomen verotuksessa luovutustappiot vähennetään ensin saman vuoden voitoista "
            "ja ylijäävä osa seuraavien 5 vuoden voitoista (ks. vero.fi/kryptovarojen-verotus).",
        ]
    )

    _autosize_columns(ws, [14, 12, 18, 18, 18, 28])
    _autosize_columns(ws_summary, [32, 48])

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    filename = f"krypto-veroraportti-{datetime.now(timezone.utc).strftime('%Y-%m-%d')}.xlsx"
    return buffer, filename
